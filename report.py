# -*- coding: utf-8 -*-
"""توليد تقرير Excel ملوّن من نتيجة المطابقة."""
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

HEAD = PatternFill("solid", fgColor="1F4E78")
HEADF = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D9D9D9")
B = Border(THIN, THIN, THIN, THIN)
COLORS = {"matched": "C6EFCE", "amount_diff": "FFD8A8", "ledger_only": "FFEB9C",
          "bank_only": "BDD7EE", "duplicates": "FFC7CE"}


def _sheet(wb, title, headers, rows, color, widths):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    fill = PatternFill("solid", fgColor=color)
    for c, h in enumerate(headers, 1):
        x = ws.cell(1, c, h); x.fill = HEAD; x.font = HEADF
        x.alignment = Alignment("center", "center"); x.border = B
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            x = ws.cell(r, c, v); x.fill = fill; x.border = B
            x.alignment = Alignment(horizontal="center")
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def build_report(result: dict) -> bytes:
    wb = openpyxl.Workbook()
    s = result["summary"]
    ws0 = wb.active; ws0.title = "الملخص"; ws0.sheet_view.rightToLeft = True
    ws0["B2"] = "تقرير المطابقة البنكية الآلية"
    ws0["B2"].font = Font(bold=True, size=15, color="1F4E78")
    rows = [("دفعات مطابقة", s["matched"], COLORS["matched"]),
            ("فرق في الإجمالي", s["amount_diff"], COLORS["amount_diff"]),
            ("في النظام فقط", s["ledger_only"], COLORS["ledger_only"]),
            ("في البنك فقط", s["bank_only"], COLORS["bank_only"]),
            ("صرف مكرر", s["duplicates"], COLORS["duplicates"])]
    for i, (lbl, n, col) in enumerate(rows, 4):
        ws0.cell(i, 2, lbl).fill = PatternFill("solid", fgColor=col)
        ws0.cell(i, 2).border = B
        ws0.cell(i, 3, n).border = B
        ws0.cell(i, 3).alignment = Alignment(horizontal="center")
    ws0.column_dimensions["B"].width = 22
    ws0.column_dimensions["C"].width = 12

    Hb = ["الرقم المرجعي", "إجمالي النظام", "إجمالي البنك", "الفرق", "البيان", "عدد السطور"]
    wd = [16, 16, 16, 12, 36, 12]
    _sheet(wb, "دفعات مطابقة", Hb,
           [(b["ref"], b["ledger_total"], b["bank_total"], b["diff"], b["description"], b["lines"])
            for b in result["matched"]], COLORS["matched"], wd)
    _sheet(wb, "فرق في الإجمالي", Hb,
           [(b["ref"], b["ledger_total"], b["bank_total"], b["diff"], b["description"], b["lines"])
            for b in result["amount_diff"]], COLORS["amount_diff"], wd)
    _sheet(wb, "في النظام فقط", Hb,
           [(b["ref"], b["ledger_total"], b["bank_total"], b["diff"], b["description"], b["lines"])
            for b in result["ledger_only"]], COLORS["ledger_only"], wd)
    _sheet(wb, "في البنك فقط", ["الرقم المرجعي", "إجمالي البنك", "البيان"],
           [(b["ref"], b["bank_total"], b["description"]) for b in result["bank_only"]],
           COLORS["bank_only"], [16, 16, 40])
    _sheet(wb, "صرف مكرر", ["التاريخ", "الرقم المرجعي", "المبلغ", "البيان", "المستفيد"],
           [(e["txn_date"] or "", e["ref"], round(abs(e["amount"]), 3), e["description"],
             e["employee"] or "") for e in result["duplicates"]], COLORS["duplicates"],
           [14, 16, 14, 34, 14])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
