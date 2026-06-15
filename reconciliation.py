# -*- coding: utf-8 -*-
"""
محرك المطابقة البنكية | Reconciliation Engine
طبقات نظيفة: قراءة → توحيد → مطابقة (على مستوى الدفعة) → نتيجة منظّمة.
"""
from __future__ import annotations
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
import openpyxl


# ───────────── نموذج موحّد ─────────────
@dataclass
class Entry:
    source: str                  # LEDGER | BANK
    ref: str | None              # الرقم المرجعي البنكي
    amount: float                # موجب=إضافة، سالب=صرف
    txn_date: str | None = None  # YYYY-MM-DD
    description: str = ""
    voucher: str | None = None
    employee: str | None = None
    row: int | None = None


def _s(v):
    if v is None:
        return None
    return (str(v).strip().replace("\xa0", "") or None)


def _ref(v):
    if v is None:
        return None
    s = str(v).strip().replace("\xa0", "").replace(" ", "")
    return s or None


def _date(v):
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    return None


# ───────────── 1) قراءة دفتر الخزينة ─────────────
def parse_ledger(path: str, sheet: str) -> list[Entry]:
    """يقرأ ورقة شهر. B تاريخ إضافة | C تاريخ صرف | D إضافة | E خصم
    F رقم المعاملة | G البيان | H رقم السند | J الرقم الوظيفي."""
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    out: list[Entry] = []
    for r in range(4, ws.max_row + 1):
        credit = ws.cell(r, 4).value
        debit = ws.cell(r, 5).value
        ref = _ref(ws.cell(r, 6).value)
        desc = str(ws.cell(r, 7).value or "").strip()
        voucher = _s(ws.cell(r, 8).value)
        emp = _s(ws.cell(r, 10).value)
        if isinstance(credit, (int, float)) and credit:
            out.append(Entry("LEDGER", ref, +float(credit),
                             _date(ws.cell(r, 2).value), desc, voucher, emp, r))
        elif isinstance(debit, (int, float)) and debit:
            out.append(Entry("LEDGER", ref, -float(debit),
                             _date(ws.cell(r, 3).value), desc, voucher, emp, r))
    return out


# ───────────── 2) قراءة كشف البنك (أعمدة مرنة) ─────────────
def parse_bank(path: str, sheet: str | None = None) -> list[Entry]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    head = {str(ws.cell(1, c).value).strip().lower(): c
            for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

    def col(*names):
        for n in names:
            if n in head:
                return head[n]
        return None

    c_ref = col("reference", "ref", "رقم المعاملة", "المرجع")
    c_date = col("date", "value date", "التاريخ", "تاريخ")
    c_amt = col("amount", "value", "المبلغ")
    c_desc = col("description", "narrative", "البيان", "details")
    out: list[Entry] = []
    for r in range(2, ws.max_row + 1):
        amt = ws.cell(r, c_amt).value if c_amt else None
        if not isinstance(amt, (int, float)):
            continue
        out.append(Entry("BANK", _ref(ws.cell(r, c_ref).value) if c_ref else None,
                         float(amt), _date(ws.cell(r, c_date).value) if c_date else None,
                         str(ws.cell(r, c_desc).value or "").strip() if c_desc else "", row=r))
    return out


# ───────────── 3) المطابقة على مستوى الدفعة ─────────────
@dataclass
class Batch:
    ref: str
    ledger_total: float = 0.0
    bank_total: float = 0.0
    lines: int = 0
    description: str = ""

    @property
    def diff(self):
        return round(self.ledger_total - self.bank_total, 3)


def reconcile(ledger: list[Entry], bank: list[Entry], tol: float = 0.01) -> dict:
    duplicates: list[Entry] = []
    seen: set[tuple] = set()
    for e in ledger:
        if e.amount < 0 and e.employee:
            ident = (e.employee, round(abs(e.amount), 3))
            if ident in seen:
                duplicates.append(e)
            else:
                seen.add(ident)

    batches: dict[str, Batch] = {}
    no_ref_bank: list[Entry] = []
    for e in ledger:
        if not e.ref:
            continue
        b = batches.setdefault(e.ref, Batch(e.ref))
        b.ledger_total += abs(e.amount)
        b.lines += 1
        if not b.description and e.amount < 0:
            b.description = e.description
    for x in bank:
        if not x.ref:
            no_ref_bank.append(x)
            continue
        b = batches.setdefault(x.ref, Batch(x.ref))
        b.bank_total += abs(x.amount)
        if not b.description:
            b.description = x.description

    matched, amount_diff, ledger_only, bank_only = [], [], [], []
    for b in batches.values():
        has_l = b.ledger_total > 0
        has_b = b.bank_total > 0
        if has_l and not has_b:
            ledger_only.append(b)
        elif has_b and not has_l:
            bank_only.append(b)
        elif abs(b.ledger_total - b.bank_total) <= tol:
            matched.append(b)
        else:
            amount_diff.append(b)
    for x in no_ref_bank:
        bank_only.append(Batch("بلا مرجع", bank_total=abs(x.amount), description=x.description))

    def tot(bs, fld):
        return round(sum(getattr(b, fld) for b in bs), 3)

    return {
        "summary": {
            "matched": len(matched), "matched_total": tot(matched, "ledger_total"),
            "amount_diff": len(amount_diff),
            "ledger_only": len(ledger_only), "ledger_only_total": tot(ledger_only, "ledger_total"),
            "bank_only": len(bank_only), "bank_only_total": tot(bank_only, "bank_total"),
            "duplicates": len(duplicates),
            "duplicates_total": round(sum(abs(e.amount) for e in duplicates), 3),
            "ledger_lines": len(ledger), "bank_lines": len(bank),
        },
        "matched": [asdict(b) | {"diff": b.diff} for b in matched],
        "amount_diff": [asdict(b) | {"diff": b.diff} for b in amount_diff],
        "ledger_only": [asdict(b) | {"diff": b.diff} for b in ledger_only],
        "bank_only": [asdict(b) | {"diff": b.diff} for b in bank_only],
        "duplicates": [asdict(e) for e in duplicates],
    }
