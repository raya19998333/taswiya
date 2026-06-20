# -*- coding: utf-8 -*-
"""
نظام التسوية البنكية الآلية — الإصدار الثاني | FastAPI Backend v2
يضيف: مصادقة + صلاحيات (RBAC) + حفظ في قاعدة بيانات + سجل تسويات + اعتماد المشرف + سجل تدقيق.
"""
import io
import re
import uuid
from collections import Counter
from shutil import which
from datetime import datetime
from pathlib import Path
from typing import Optional
import openpyxl
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas as _rl_canvas

from reconciliation import parse_ledger, parse_bank, reconcile
from report import build_report
from database import init_db, get_conn, log_action
from auth import login, current_user, require_role

try:
    from PIL import Image
except Exception:  # pragma: no cover - optional dependency at runtime
    Image = None

try:
    import pytesseract
except Exception:  # pragma: no cover - optional dependency at runtime
    pytesseract = None

try:
    import fitz  # PyMuPDF
except Exception:  # pragma: no cover - optional dependency at runtime
    fitz = None

if pytesseract is not None:
    detected_tesseract = which("tesseract")
    if not detected_tesseract and Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe").exists():
        detected_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if not detected_tesseract and Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe").exists():
        detected_tesseract = r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"
    if detected_tesseract:
        pytesseract.pytesseract.tesseract_cmd = detected_tesseract

app = FastAPI(title="نظام التسوية البنكية الآلية")
BASE = Path(__file__).parent
STATIC_DIR = BASE / "static"
UPLOAD_DIR = BASE / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
init_db()

# Mount static files
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


def _read(upload: UploadFile, parser, sheet=None):
    try:
        tmp = io.BytesIO(upload.file.read())
        return parser(tmp, sheet) if sheet is not None else parser(tmp)
    except Exception as e:
        raise HTTPException(400, f"تعذّر قراءة الملف {upload.filename}: {e}")


def _file_kind(filename: str) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".bmp"}:
        return "image"
    if suffix == ".pdf":
        return "pdf"
    return "other"


def _file_kind_from_bytes(raw: bytes, filename: str = "", content_type: str = "") -> str:
    kind = _file_kind(filename)
    if kind != "other":
        return kind

    ctype = (content_type or "").lower()
    if ctype.startswith("image/"):
        return "image"
    if ctype == "application/pdf":
        return "pdf"

    if raw.startswith(b"%PDF"):
        return "pdf"

    try:
        if Image is not None:
            Image.open(io.BytesIO(raw)).verify()
            return "image"
    except Exception:
        pass
    return "other"


def _normalize_amount(value: str):
    if not value:
        return None
    value = value.replace(",", "").strip()
    try:
        return float(value)
    except ValueError:
        return None


def _next_voucher_no(conn, voucher_date: str | None = None) -> str:
    year = (voucher_date or datetime.utcnow().date().isoformat())[:4]
    prefix = f"V-{year}-"
    rows = conn.execute(
        "SELECT voucher_no FROM vouchers WHERE voucher_no LIKE ? ORDER BY id DESC LIMIT 200",
        (f"{prefix}%",),
    ).fetchall()
    max_seq = 0
    for row in rows:
        value = row[0] or ""
        match = re.search(rf"{re.escape(prefix)}(\d+)$", value)
        if match:
            max_seq = max(max_seq, int(match.group(1)))
    return f"{prefix}{max_seq + 1:04d}"


def _voucher_filters_sql(q=None, date_from=None, date_to=None, amount_min=None, amount_max=None, status=None):
    sql = """FROM vouchers v
             JOIN users u ON u.id=v.created_by WHERE 1=1"""
    params = []
    if q:
        sql += " AND (v.voucher_no LIKE ? OR v.beneficiary LIKE ? OR v.source_unit LIKE ? OR v.description LIKE ? OR v.ocr_text LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like, like]
    if date_from:
        sql += " AND v.voucher_date >= ?"; params.append(date_from)
    if date_to:
        sql += " AND v.voucher_date <= ?"; params.append(date_to)
    if amount_min is not None:
        sql += " AND v.amount >= ?"; params.append(amount_min)
    if amount_max is not None:
        sql += " AND v.amount <= ?"; params.append(amount_max)
    if status:
        sql += " AND v.status = ?"; params.append(status)
    return sql, params


def _ocr_image_bytes(raw: bytes) -> str:
    if Image is None or pytesseract is None:
        raise HTTPException(503, "ميزة المسح التلقائي غير متاحة حالياً على الخادم")
    image = Image.open(io.BytesIO(raw)).convert("L")
    return pytesseract.image_to_string(image, lang="ara+eng").strip()


def _extract_text_from_pdf(raw: bytes) -> str:
    if fitz is None:
        raise HTTPException(503, "ميزة قراءة ملفات PDF غير متاحة حالياً على الخادم")

    doc = fitz.open(stream=raw, filetype="pdf")
    pieces = []
    try:
        for page_index, page in enumerate(doc):
            text = page.get_text("text").strip()
            if text:
                pieces.append(text)
            if len(pieces) >= 3 and sum(len(part) for part in pieces) > 120:
                break

        extracted = "\n".join(pieces).strip()
        if extracted:
            return extracted

        if Image is None or pytesseract is None:
            raise HTTPException(503, "ميزة المسح التلقائي غير متاحة حالياً على الخادم")

        for page_index in range(min(len(doc), 3)):
            page = doc[page_index]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("L")
            text = pytesseract.image_to_string(img, lang="ara+eng").strip()
            if text:
                pieces.append(text)
            if len("\n".join(pieces)) > 200:
                break
        return "\n".join(pieces).strip()
    finally:
        doc.close()


def _extract_voucher_data(ocr_text: str) -> dict:
    text = (ocr_text or "").strip()
    compact = re.sub(r"\s+", " ", text)
    voucher_no = None
    voucher_date = None
    amount = None
    beneficiary = None
    source_unit = None

    patterns = [
        r"(?:رقم\s*السند|رقم|voucher)\s*[:#-]?\s*([A-Za-z0-9\-\/]+)",
        r"(?:السند)\s*[:#-]?\s*([A-Za-z0-9\-\/]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            voucher_no = match.group(1).strip()
            break

    match = re.search(r"(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4})", compact)
    if match:
        voucher_date = match.group(1).replace("/", "-")

    for pattern in (
        r"(?:المبلغ|amount|total|الإجمالي)\s*[:=]?\s*([0-9][0-9,]*(?:\.[0-9]+)?)",
        r"([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:ر\.ع|ريال|omr)",
    ):
        match = re.search(pattern, compact, re.I)
        if match:
            amount = _normalize_amount(match.group(1))
            if amount is not None:
                break

    for label in ("المستفيد", "beneficiary", "payee", "إلى", "صرف إلى"):
        match = re.search(rf"{label}\s*[:=]?\s*([^\n\r|]+)", text, re.I)
        if match:
            beneficiary = match.group(1).strip()
            break

    for label in ("الخزينة", "الجهة", "الادارة", "الإدارة", "القسم", "unit", "department"):
        match = re.search(rf"{label}\s*[:=]?\s*([^\n\r|]+)", text, re.I)
        if match:
            source_unit = match.group(1).strip()
            break

    return {
        "voucher_no": voucher_no,
        "voucher_date": voucher_date,
        "amount": amount,
        "beneficiary": beneficiary,
        "source_unit": source_unit,
        "description": compact[:300],
        "ocr_text": text,
    }


def _is_readable_document(ocr_text: str, parsed: dict) -> bool:
    text = re.sub(r"\s+", "", (ocr_text or "")).strip()
    if len(text) < 20:
        return False

    filled_fields = sum(
        1
        for key in ("voucher_no", "voucher_date", "amount", "beneficiary", "source_unit")
        if parsed.get(key) not in (None, "")
    )
    return len(text) >= 30 or filled_fields >= 2


async def _read_voucher_ocr(file: UploadFile, raw: bytes | None = None) -> tuple[str, str]:
    if not file or not file.filename:
        return "", ""
    raw = raw if raw is not None else await file.read()
    try:
        kind = _file_kind_from_bytes(raw, file.filename, getattr(file, "content_type", ""))
        if kind == "image":
            return _ocr_image_bytes(raw), "tesseract"
        if kind == "pdf":
            return _extract_text_from_pdf(raw), "pdf+ocr"
        return "", ""
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"تعذّر قراءة ملف السند: {e}")


# ───────────── الواجهة ─────────────
@app.get("/", response_class=HTMLResponse)
def home():
    login_file = STATIC_DIR / "login.html"
    if login_file.exists():
        return login_file.read_text(encoding="utf-8")
    return "<h1>Error: login.html not found</h1>"

@app.get("/app", response_class=HTMLResponse)
def app_page():
    index_file = STATIC_DIR / "index.html"
    if index_file.exists():
        return index_file.read_text(encoding="utf-8")
    return "<h1>Error: index.html not found</h1>"


# ───────────── المصادقة ─────────────
@app.post("/login")
async def do_login(username: str = Form(...), password: str = Form(...)):
    token, u = login(username, password)
    log_action(u["id"], u["username"], "login", "تسجيل دخول")
    return {"token": token, "user": {"username": u["username"], "role": u["role"],
                                     "full_name": u["full_name"]}}


@app.get("/me")
async def me(user: dict = Depends(current_user)):
    return {
        "username": user["username"],
        "role": user["role"],
        "full_name": user["full_name"],
        "id": user["id"],
    }


# ───────────── أوراق الملف ─────────────
@app.post("/sheets")
async def sheets(file: UploadFile = File(...), user: dict = Depends(current_user)):
    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True)
    return {"sheets": wb.sheetnames}


# ───────────── تشغيل المطابقة + الحفظ ─────────────
@app.post("/reconcile")
async def do_reconcile(
    ledger: UploadFile = File(...), bank: UploadFile = File(...),
    ledger_sheet: str = Form(...), bank_sheet: str = Form(None),
    user: dict = Depends(current_user),
):
    led = _read(ledger, parse_ledger, ledger_sheet)
    bnk = _read(bank, parse_bank, bank_sheet or None)
    if not led:
        raise HTTPException(400, "لا توجد قيود في ورقة الخزينة المختارة.")
    result = reconcile(led, bnk)
    run_id = _save_run(user, ledger.filename, bank.filename, ledger_sheet, result)
    log_action(user["id"], user["username"], "reconcile",
               f"تشغيل تسوية #{run_id} ({ledger_sheet})")
    result["run_id"] = run_id
    result["status"] = "draft"
    return JSONResponse(result)


def _save_run(user, ledger_name, bank_name, sheet, result) -> int:
    s = result["summary"]
    conn = get_conn()
    cur = conn.execute(
        """INSERT INTO runs(created_by,created_at,ledger_name,bank_name,sheet,
           matched,amount_diff,ledger_only,bank_only,duplicates,matched_total,status)
           VALUES(?,?,?,?,?,?,?,?,?,?,?, 'draft')""",
        (user["id"], datetime.utcnow().isoformat(), ledger_name, bank_name, sheet,
         s["matched"], s["amount_diff"], s["ledger_only"], s["bank_only"],
         s["duplicates"], s["matched_total"]))
    run_id = cur.lastrowid
    for cat in ("matched", "amount_diff", "ledger_only", "bank_only"):
        for b in result[cat]:
            conn.execute(
                """INSERT INTO run_items(run_id,category,ref,ledger_total,bank_total,diff,description)
                   VALUES(?,?,?,?,?,?,?)""",
                (run_id, cat, b["ref"], b["ledger_total"], b["bank_total"], b["diff"], b["description"]))
    for e in result["duplicates"]:
        conn.execute(
            """INSERT INTO run_items(run_id,category,ref,description,employee,txn_date,amount)
               VALUES(?,?,?,?,?,?,?)""",
            (run_id, "duplicates", e["ref"], e["description"], e["employee"],
             e["txn_date"], e["amount"]))
    conn.commit()
    conn.close()
    return run_id


# ───────────── سجل التسويات ─────────────
@app.get("/runs")
async def list_runs(user: dict = Depends(current_user)):
    conn = get_conn()
    rows = conn.execute(
        """SELECT r.*, u.full_name creator, a.full_name approver
           FROM runs r JOIN users u ON u.id=r.created_by
           LEFT JOIN users a ON a.id=r.approved_by
           ORDER BY r.id DESC LIMIT 100""").fetchall()
    conn.close()
    return {"runs": [dict(x) for x in rows]}


@app.get("/runs/{run_id}")
async def get_run(run_id: int, user: dict = Depends(current_user)):
    conn = get_conn()
    run = conn.execute("SELECT * FROM runs WHERE id=?", (run_id,)).fetchone()
    if not run:
        conn.close(); raise HTTPException(404, "التسوية غير موجودة")
    items = conn.execute("SELECT * FROM run_items WHERE run_id=?", (run_id,)).fetchall()
    conn.close()
    grouped = {}
    for it in items:
        grouped.setdefault(it["category"], []).append(dict(it))
    return {"run": dict(run), "items": grouped}


# ───────────── اعتماد المشرف (RBAC) ─────────────
@app.post("/runs/{run_id}/approve")
async def approve(run_id: int, user: dict = Depends(require_role("supervisor"))):
    conn = get_conn()
    run = conn.execute("SELECT * FROM runs WHERE id=?", (run_id,)).fetchone()
    if not run:
        conn.close(); raise HTTPException(404, "التسوية غير موجودة")
    if run["status"] == "approved":
        conn.close(); raise HTTPException(400, "التسوية معتمدة مسبقاً")
    conn.execute("UPDATE runs SET status='approved', approved_by=?, approved_at=? WHERE id=?",
                 (user["id"], datetime.utcnow().isoformat(), run_id))
    conn.commit(); conn.close()
    log_action(user["id"], user["username"], "approve", f"اعتماد التسوية #{run_id}")
    return {"ok": True, "status": "approved"}


@app.delete("/runs/{run_id}")
async def delete_run(run_id: int, user: dict = Depends(require_role("admin"))):
    conn = get_conn()
    run = conn.execute("SELECT * FROM runs WHERE id=?", (run_id,)).fetchone()
    if not run:
        conn.close(); raise HTTPException(404, "التسوية غير موجودة")
    conn.execute("DELETE FROM runs WHERE id=?", (run_id,))
    conn.commit(); conn.close()
    log_action(user["id"], user["username"], "delete", f"حذف التسوية #{run_id}")
    return {"ok": True}


@app.patch("/runs/{run_id}")
async def update_run(run_id: int, payload: dict, user: dict = Depends(require_role("admin"))):
    allowed = {"ledger_name", "bank_name", "sheet", "status"}
    data = {k: v for k, v in payload.items() if k in allowed and v is not None}
    if not data:
        raise HTTPException(400, "لا توجد بيانات للتعديل")

    conn = get_conn()
    run = conn.execute("SELECT * FROM runs WHERE id=?", (run_id,)).fetchone()
    if not run:
        conn.close(); raise HTTPException(404, "التسوية غير موجودة")

    updates = []
    params = []
    if "ledger_name" in data:
        updates.append("ledger_name = ?")
        params.append(data["ledger_name"])
    if "bank_name" in data:
        updates.append("bank_name = ?")
        params.append(data["bank_name"])
    if "sheet" in data:
        updates.append("sheet = ?")
        params.append(data["sheet"])
    if "status" in data:
        new_status = data["status"].lower()
        if new_status not in {"draft", "approved"}:
            conn.close(); raise HTTPException(400, "الحالة غير صالحة")
        if new_status == "approved":
            updates.append("status = 'approved'")
            updates.append("approved_by = ?")
            updates.append("approved_at = ?")
            params.extend((user["id"], datetime.utcnow().isoformat()))
        else:
            updates.append("status = 'draft'")
            updates.append("approved_by = NULL")
            updates.append("approved_at = NULL")
    if not updates:
        conn.close(); raise HTTPException(400, "لا توجد حقول صالحة للتعديل")

    query = f"UPDATE runs SET {', '.join(updates)} WHERE id=?"
    params.append(run_id)
    conn.execute(query, tuple(params))
    conn.commit(); conn.close()
    log_action(user["id"], user["username"], "update", f"تعديل التسوية #{run_id}")
    return {"ok": True}


# ───────────── سجل التدقيق (admin) ─────────────
@app.get("/audit")
async def audit(user: dict = Depends(require_role("admin"))):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 500").fetchall()
    conn.close()
    return {"audit": [dict(x) for x in rows]}


@app.delete("/audit/{audit_id}")
async def delete_audit_entry(audit_id: int, user: dict = Depends(require_role("admin"))):
    conn = get_conn()
    row = conn.execute("SELECT id FROM audit_log WHERE id=?", (audit_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "السجل غير موجود")
    conn.execute("DELETE FROM audit_log WHERE id=?", (audit_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/audit")
async def clear_audit_log(user: dict = Depends(require_role("admin"))):
    conn = get_conn()
    conn.execute("DELETE FROM audit_log")
    conn.commit()
    conn.close()
    return {"ok": True}


# ───────────── إدارة المستخدمين (admin) ─────────────
@app.get("/users")
async def list_users(user: dict = Depends(require_role("admin"))):
    conn = get_conn()
    rows = conn.execute(
        "SELECT id,username,role,full_name,created_at FROM users ORDER BY id"
    ).fetchall()
    conn.close()
    return {"users": [dict(x) for x in rows]}


@app.post("/users")
async def create_user(
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    full_name: str = Form(...),
    user: dict = Depends(require_role("admin")),
):
    from auth import hash_password as hp
    if role not in ("admin", "supervisor", "entry"):
        raise HTTPException(400, "دور غير صالح")
    if len(password) < 6:
        raise HTTPException(400, "كلمة المرور يجب أن تكون 6 أحرف على الأقل")
    conn = get_conn()
    if conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        conn.close()
        raise HTTPException(400, "اسم المستخدم موجود مسبقاً")
    h, s = hp(password)
    now = datetime.utcnow().isoformat()
    conn.execute(
        "INSERT INTO users(username,password_hash,salt,role,full_name,created_at)"
        " VALUES(?,?,?,?,?,?)",
        (username, h, s, role, full_name, now),
    )
    conn.commit()
    conn.close()
    log_action(user["id"], user["username"], "create_user", f"إنشاء مستخدم: {username}")
    return {"ok": True}


@app.patch("/users/{uid}")
async def update_user(uid: int, payload: dict, user: dict = Depends(require_role("admin"))):
    allowed = {"role", "full_name"}
    data = {k: v for k, v in payload.items() if k in allowed and v is not None}
    if not data:
        raise HTTPException(400, "لا توجد بيانات للتعديل")
    conn = get_conn()
    if not conn.execute("SELECT id FROM users WHERE id=?", (uid,)).fetchone():
        conn.close()
        raise HTTPException(404, "المستخدم غير موجود")
    sets = ", ".join(f"{k}=?" for k in data)
    conn.execute(f"UPDATE users SET {sets} WHERE id=?", [*data.values(), uid])
    conn.commit()
    conn.close()
    log_action(user["id"], user["username"], "update_user", f"تعديل مستخدم #{uid}")
    return {"ok": True}


@app.delete("/users/{uid}")
async def delete_user(uid: int, user: dict = Depends(require_role("admin"))):
    if uid == user["id"]:
        raise HTTPException(400, "لا يمكن حذف حسابك الخاص")
    conn = get_conn()
    if not conn.execute("SELECT id FROM users WHERE id=?", (uid,)).fetchone():
        conn.close()
        raise HTTPException(404, "المستخدم غير موجود")
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    log_action(user["id"], user["username"], "delete_user", f"حذف مستخدم #{uid}")
    return {"ok": True}


@app.post("/users/{uid}/reset-password")
async def reset_user_password(uid: int, payload: dict, user: dict = Depends(require_role("admin"))):
    from auth import hash_password as hp
    new_pass = (payload.get("password") or "").strip()
    if len(new_pass) < 6:
        raise HTTPException(400, "كلمة المرور يجب أن تكون 6 أحرف على الأقل")
    conn = get_conn()
    if not conn.execute("SELECT id FROM users WHERE id=?", (uid,)).fetchone():
        conn.close()
        raise HTTPException(404, "المستخدم غير موجود")
    h, s = hp(new_pass)
    conn.execute(
        "UPDATE users SET password_hash=?, salt=?, password_changed_at=? WHERE id=?",
        (h, s, datetime.utcnow().isoformat(), uid),
    )
    conn.commit()
    conn.close()
    log_action(user["id"], user["username"], "reset_password", f"إعادة ضبط كلمة مرور #{uid}")
    return {"ok": True}


# ───────────── سندات الصرف (الأرشيف) ─────────────
@app.post("/vouchers")
async def create_voucher(
    voucher_no: str = Form(None),
    voucher_date: str = Form(...),
    amount: float = Form(...),
    currency: str = Form("ر.ع"),
    source_unit: str = Form(None),
    beneficiary: str = Form(...),
    description: str = Form(None),
    ocr_text: str = Form(None),
    ocr_source: str = Form(None),
    file: Optional[UploadFile] = File(None),
    user: dict = Depends(current_user),
):
    file_path = None
    file_name = None
    extracted_ocr_text = ocr_text
    extracted_ocr_source = ocr_source
    if file and file.filename:
        contents = await file.read()
        kind = _file_kind_from_bytes(contents, file.filename, getattr(file, "content_type", ""))
        if kind not in {"image", "pdf"}:
            raise HTTPException(400, "صيغة الملف غير مدعومة (PDF أو صورة فقط)")
        ext = Path(file.filename).suffix.lower()
        if not ext:
            ext = ".pdf" if kind == "pdf" else ".png"
        safe_name = f"{uuid.uuid4().hex}{ext}"
        (UPLOAD_DIR / safe_name).write_bytes(contents)
        file_path = safe_name
        file_name = file.filename

        extracted_text = None
        extracted_source = None
        try:
            extracted_text, extracted_source = await _read_voucher_ocr(file, contents)
        except HTTPException:
            raise
        except Exception:
            extracted_text = None
            extracted_source = None

        if extracted_text:
            parsed = _extract_voucher_data(extracted_text)
            voucher_no = voucher_no or parsed["voucher_no"]
            voucher_date = voucher_date or parsed["voucher_date"] or voucher_date
            amount = amount if amount is not None else parsed["amount"]
            source_unit = source_unit or parsed["source_unit"]
            beneficiary = beneficiary or parsed["beneficiary"]
            description = description or parsed["description"]
            extracted_ocr_text = extracted_ocr_text or extracted_text
            extracted_ocr_source = extracted_ocr_source or extracted_source

    conn = get_conn()
    if not voucher_no:
        voucher_no = _next_voucher_no(conn, voucher_date)

    cur = conn.execute(
        """INSERT INTO vouchers(voucher_no,voucher_date,amount,currency,source_unit,beneficiary,
           description,ocr_text,ocr_source,file_path,file_name,created_by,created_at)
           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (voucher_no, voucher_date, amount, currency, source_unit, beneficiary, description,
         extracted_ocr_text, extracted_ocr_source, file_path, file_name, user["id"], datetime.utcnow().isoformat()))
    voucher_id = cur.lastrowid
    conn.commit(); conn.close()
    log_action(user["id"], user["username"], "voucher_create", f"إضافة سند #{voucher_id}")
    return {"ok": True, "id": voucher_id}


@app.post("/vouchers/ocr")
async def voucher_ocr(file: UploadFile = File(...), user: dict = Depends(current_user)):
    if not file.filename:
        raise HTTPException(400, "يرجى رفع ملف صورة أو PDF")

    try:
        raw = await file.read()
        kind = _file_kind_from_bytes(raw, file.filename, getattr(file, "content_type", ""))
        if kind not in {"image", "pdf"}:
            raise HTTPException(400, "يرجى رفع صورة أو PDF")
        ocr_text, ocr_source = await _read_voucher_ocr(file, raw)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"تعذّر قراءة ملف السند: {e}")

    parsed = _extract_voucher_data(ocr_text)
    if not _is_readable_document(ocr_text, parsed):
        raise HTTPException(422, "تعذّر قراءة المستند — يرجى رفع نسخة أوضح.")

    parsed.update({"ocr_text": ocr_text, "ocr_source": ocr_source})
    return parsed


@app.get("/vouchers")
async def list_vouchers(
    q: str = None, date_from: str = None, date_to: str = None,
    amount_min: float = None, amount_max: float = None, status: str = None,
    user: dict = Depends(current_user),
):
    conn = get_conn()
    sql, params = _voucher_filters_sql(q, date_from, date_to, amount_min, amount_max, status)
    sql = "SELECT v.*, u.full_name creator " + sql
    sql += " ORDER BY v.id DESC LIMIT 200"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return {"vouchers": [dict(x) for x in rows]}


@app.get("/vouchers/summary")
async def vouchers_summary(
    q: str = None, date_from: str = None, date_to: str = None,
    amount_min: float = None, amount_max: float = None, status: str = None,
    user: dict = Depends(current_user),
):
    conn = get_conn()
    sql, params = _voucher_filters_sql(q, date_from, date_to, amount_min, amount_max, status)
    rows = conn.execute("SELECT v.status, v.amount " + sql, params).fetchall()
    recent = conn.execute(
        "SELECT v.*, u.full_name creator FROM vouchers v JOIN users u ON u.id=v.created_by ORDER BY v.id DESC LIMIT 5"
    ).fetchall()
    conn.close()
    status_counts = Counter(row[0] for row in rows)
    total_amount = sum(float(row[1] or 0) for row in rows)
    return {
        "counts": {"draft": status_counts.get("draft", 0), "approved": status_counts.get("approved", 0), "cancelled": status_counts.get("cancelled", 0)},
        "total_amount": total_amount,
        "recent": [dict(x) for x in recent],
    }


def _build_voucher_report_rows(rows):
    return [[
        r["voucher_no"] or "",
        r["voucher_date"] or "",
        r["source_unit"] or "",
        r["beneficiary"] or "",
        f"{float(r['amount'] or 0):.3f}",
        r["currency"] or "",
        r["status"] or "",
        r["description"] or "",
    ] for r in rows]


@app.get("/vouchers/report")
async def vouchers_report(
    format: str = "excel",
    q: str = None, date_from: str = None, date_to: str = None,
    amount_min: float = None, amount_max: float = None, status: str = None,
    user: dict = Depends(current_user),
):
    conn = get_conn()
    sql, params = _voucher_filters_sql(q, date_from, date_to, amount_min, amount_max, status)
    rows = conn.execute("SELECT v.* " + sql + " ORDER BY v.id DESC", params).fetchall()
    conn.close()

    headers = ["رقم السند", "التاريخ", "الجهة/الخزينة", "المستفيد", "المبلغ", "العملة", "الحالة", "البيان"]
    data_rows = _build_voucher_report_rows(rows)

    if format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vouchers"
        ws.append(headers)
        for row in data_rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        log_action(user["id"], user["username"], "voucher_report", "تصدير تقرير سندات Excel")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=vouchers_report.xlsx"},
        )

    if format == "pdf":
        navy    = colors.HexColor("#1e3a5f")
        navy2   = colors.HexColor("#2d5a8e")
        teal    = colors.HexColor("#2bb5ae")
        bg2     = colors.HexColor("#eef1f6")
        bdr     = colors.HexColor("#d0dae8")
        muted   = colors.HexColor("#6b7d93")
        c_green = colors.HexColor("#28c17b")
        c_amber = colors.HexColor("#f0a43a")
        PAGE    = landscape(A4)
        pw      = PAGE[0]

        class _NC(_rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                _rl_canvas.Canvas.__init__(self, *args, **kwargs)
                self._pg = []
            def showPage(self):
                self._pg.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                n = len(self._pg)
                for i, p in enumerate(self._pg, 1):
                    self.__dict__.update(p)
                    self._footer(i, n)
                    _rl_canvas.Canvas.showPage(self)
                _rl_canvas.Canvas.save(self)
            def _footer(self, num, total):
                self.saveState()
                self.setStrokeColor(bdr)
                self.setLineWidth(0.5)
                self.line(15*mm, 14*mm, pw - 15*mm, 14*mm)
                self.setFont("Helvetica", 7.5)
                self.setFillColor(muted)
                self.drawString(15*mm, 9*mm, "Automated Bank Reconciliation System")
                self.drawCentredString(pw / 2, 9*mm,
                                       datetime.utcnow().strftime("Exported: %Y-%m-%d %H:%M UTC"))
                self.drawRightString(pw - 15*mm, 9*mm, f"Page {num} / {total}")
                self.restoreState()

        def mk(name, **kw):
            s = ParagraphStyle(name)
            s.fontName  = kw.get("f", "Helvetica")
            s.fontSize  = kw.get("sz", 9)
            s.textColor = kw.get("c", colors.black)
            s.alignment = kw.get("a", 0)
            s.leading   = kw.get("sz", 9) + 3
            return s

        total_cnt = len(rows)
        total_amt = sum(float(r["amount"] or 0) for r in rows)
        approved  = sum(1 for r in rows if (r["status"] or "") == "approved")
        pending   = total_cnt - approved

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=PAGE,
                                rightMargin=15*mm, leftMargin=15*mm,
                                topMargin=15*mm, bottomMargin=22*mm)
        cw  = pw - 30*mm
        cw4 = cw / 4
        story = []

        hdr = Table([
            [Paragraph("Bank Reconciliation System",
                       mk("ph1", f="Helvetica-Bold", sz=16, c=colors.white, a=1))],
            [Paragraph("Voucher Report",
                       mk("ph2", f="Helvetica-Bold", sz=11, c=teal, a=1))],
            [Paragraph(datetime.utcnow().strftime("Generated: %B %d, %Y"),
                       mk("ph3", sz=8, c=colors.HexColor("#90aac8"), a=1))],
        ], colWidths=[cw])
        hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), navy),
            ("TOPPADDING",    (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING",   (0, 0), (-1, -1), 20),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 20),
            ("LINEBELOW",     (0, -1), (-1, -1), 3, teal),
        ]))
        story += [hdr, Spacer(1, 5*mm)]

        stats = Table([
            [Paragraph(str(total_cnt), mk("ps1", f="Helvetica-Bold", sz=16, c=navy,    a=1)),
             Paragraph(str(approved),  mk("ps2", f="Helvetica-Bold", sz=16, c=c_green, a=1)),
             Paragraph(str(pending),   mk("ps3", f="Helvetica-Bold", sz=16, c=c_amber, a=1)),
             Paragraph(f"{total_amt:,.3f}", mk("ps4", f="Helvetica-Bold", sz=14, c=teal, a=1))],
            [Paragraph("Total Vouchers",     mk("pl1", sz=7.5, c=muted, a=1)),
             Paragraph("Approved",           mk("pl2", sz=7.5, c=muted, a=1)),
             Paragraph("Pending / Draft",    mk("pl3", sz=7.5, c=muted, a=1)),
             Paragraph("Total Amount (OMR)", mk("pl4", sz=7.5, c=muted, a=1))],
        ], colWidths=[cw4] * 4)
        stats.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, -1), colors.HexColor("#e8f0fa")),
            ("BACKGROUND",    (1, 0), (1, -1), colors.HexColor("#e6f5ea")),
            ("BACKGROUND",    (2, 0), (2, -1), colors.HexColor("#fff8e6")),
            ("BACKGROUND",    (3, 0), (3, -1), colors.HexColor("#e6f9f7")),
            ("BOX",           (0, 0), (-1, -1), 0.5, bdr),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, bdr),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story += [stats, Spacer(1, 5*mm)]

        cws  = [25*mm, 22*mm, 30*mm, 45*mm, 22*mm, 14*mm, 17*mm, None]
        h_s  = mk("pth", f="Helvetica-Bold", sz=8, c=colors.white, a=1)
        d_s  = mk("ptd", sz=7.5, c=colors.HexColor("#1e2d3d"), a=0)
        hdrs = ["Voucher No.", "Date", "Source Unit", "Beneficiary",
                "Amount (OMR)", "Currency", "Status", "Description"]
        tdata = [[Paragraph(h, h_s) for h in hdrs]]
        for dr in data_rows:
            tdata.append([Paragraph(str(v), d_s) for v in dr])
        mtbl = Table(tdata, colWidths=cws, repeatRows=1)
        mtbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0), navy),
            ("LINEBELOW",      (0, 0), (-1, 0),  2, teal),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg2]),
            ("GRID",  (0, 0), (-1, -1), 0.3, bdr),
            ("BOX",   (0, 0), (-1, -1), 1,   navy2),
            ("TOPPADDING",    (0, 0), (-1, 0),  8),
            ("BOTTOMPADDING", (0, 0), (-1, 0),  8),
            ("TOPPADDING",    (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(mtbl)

        doc.build(story, canvasmaker=_NC)
        buf.seek(0)
        log_action(user["id"], user["username"], "voucher_report", "تصدير تقرير سندات PDF")
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=vouchers_report.pdf"},
        )

    raise HTTPException(400, "صيغة التقرير غير مدعومة")


@app.get("/vouchers/{voucher_id}/file")
async def get_voucher_file(voucher_id: int, user: dict = Depends(current_user)):
    conn = get_conn()
    v = conn.execute("SELECT * FROM vouchers WHERE id=?", (voucher_id,)).fetchone()
    conn.close()
    if not v or not v["file_path"]:
        raise HTTPException(404, "لا يوجد ملف مرفق لهذا السند")
    path = UPLOAD_DIR / v["file_path"]
    if not path.exists():
        raise HTTPException(404, "الملف غير موجود على الخادم")
    return FileResponse(path, filename=v["file_name"])


@app.patch("/vouchers/{voucher_id}")
async def update_voucher_status(
    voucher_id: int, payload: dict, user: dict = Depends(require_role("supervisor"))
):
    new_status = (payload or {}).get("status")
    if new_status not in {"draft", "approved", "cancelled"}:
        raise HTTPException(400, "حالة غير صالحة")
    conn = get_conn()
    v = conn.execute("SELECT * FROM vouchers WHERE id=?", (voucher_id,)).fetchone()
    if not v:
        conn.close(); raise HTTPException(404, "السند غير موجود")
    conn.execute("UPDATE vouchers SET status=? WHERE id=?", (new_status, voucher_id))
    conn.commit(); conn.close()
    log_action(user["id"], user["username"], "voucher_update",
               f"تحديث حالة السند #{voucher_id} إلى {new_status}")
    return {"ok": True}


@app.delete("/vouchers/{voucher_id}")
async def delete_voucher(voucher_id: int, user: dict = Depends(require_role("admin"))):
    conn = get_conn()
    v = conn.execute("SELECT * FROM vouchers WHERE id=?", (voucher_id,)).fetchone()
    if not v:
        conn.close(); raise HTTPException(404, "السند غير موجود")
    conn.execute("DELETE FROM vouchers WHERE id=?", (voucher_id,))
    conn.commit(); conn.close()
    if v["file_path"]:
        (UPLOAD_DIR / v["file_path"]).unlink(missing_ok=True)
    log_action(user["id"], user["username"], "voucher_delete", f"حذف السند #{voucher_id}")
    return {"ok": True}


# ───────────── تقرير Excel ─────────────
@app.post("/report")
async def report(
    ledger: UploadFile = File(...), bank: UploadFile = File(...),
    ledger_sheet: str = Form(...), bank_sheet: str = Form(None),
    user: dict = Depends(current_user),
):
    led = _read(ledger, parse_ledger, ledger_sheet)
    bnk = _read(bank, parse_bank, bank_sheet or None)
    data = build_report(reconcile(led, bnk))
    log_action(user["id"], user["username"], "export", "تصدير تقرير Excel")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reconciliation_report.xlsx"})
